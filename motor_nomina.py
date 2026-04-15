"""
Sistema de Nómina Automatizada — GRANDA VARGAS SAS / Soccer 7
Genera exactamente: REPORTE_HORARIOS_YYYYMMDD_YYYYMMDD.xlsx
                    COLILLA_DE_PAGO_YYYYMMDD_YYYYMMDD.xlsx

Uso:
    python nomina_soccer7.py reporte_reloj.csv
    python nomina_soccer7.py reporte_reloj.csv --inicio 2026-04-01 --fin 2026-04-15

Fórmulas confirmadas contra archivos originales:
    valor_hora       = salario / 220
    hora_extra       = valor_hora × 1.25
    recargo_nocturno = valor_hora × 0.35
    quincena         = 88h = 3.6667 días (fracción)
    max extras nóm.  = 8h
    IBC cotización   = (salario/2) + auxilio_extras + recargo_noct
    pension/salud    = IBC × 4% cada una
    auxilio transp.  = 249,095 / 2 = 124,547.50 por quincena
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import argparse, calendar

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN — EDITAR AQUÍ
# ═══════════════════════════════════════════════════════════════════════════════

EMPRESA = {
    "nombre": "GRANDA VARGAS SAS",
    "nit":    "901436091-0",
    "dir":    "CRA 51 45 -45",
    "tel":    "8474747",
}

AUXILIO_TRANSPORTE_ANUAL = 249095   # 2026
HORAS_MES       = 220               # divisor valor hora Colombia
HORAS_QUINCENA  = 88
MAX_EXTRAS_NOM  = 8                 # máx horas extras en nómina
HORA_NOCT_INI   = 19                # desde 19:00 = recargo nocturno
HORA_MADRUGADA  = 6                 # salidas antes de 6am = día anterior
FACTOR_EXTRA    = 1.25
FACTOR_NOCT     = 0.35
PENSION_PCT     = 0.04
SALUD_PCT       = 0.04
SALARIO_MINIMO  = 1423500           # 2026

# Datos de cada colaborador: (cedula, nombre_completo, cargo, banco, cuenta, eps, salario)
COLABORADORES = {
    # clave: (cedula, nombre_completo, cargo, banco, num_cuenta, eps, salario_mensual, tipo)
    # tipo: "empleado" = contrato laboral (deducciones salud/pension)
    #       "prestador" = prestador de servicio (sin deducciones, pago por hora)
    "BERTHA RESTREPO":      ("43413529",   "BERTHA LIBIA RESTREPO LEDEZMA",      "JEFE DE COCINA",           "AHORROS DAVIVIENDA", "3974-0007-7277", "NUEVA EPS",   1806565, "empleado"),
    "ERIKA  BERMUDEZ":      ("1047994257", "ERICA YORLADIS BERMUDEZ RAMIREZ",    "AUXILIAR DE COCINA",       "AHORROS DAVIVIENDA", "4884-4268-0267", "SAVIA SALUD", 1750905, "empleado"),
    "ERIKA BERMUDEZ":       ("1047994257", "ERICA YORLADIS BERMUDEZ RAMIREZ",    "AUXILIAR DE COCINA",       "AHORROS DAVIVIENDA", "4884-4268-0267", "SAVIA SALUD", 1750905, "empleado"),
    "DANIELA SANCHEZ":      ("1033340824", "DANIELA SANCHEZ TORRES",             "AUXILIAR DE COCINA",       "AHORROS DAVIVIENDA", "4884-3737-2466", "SAVIA SALUD", 1750905, "empleado"),
    "CAROLINA ARANGO":      ("1033336422", "CAROLINA ARANGO GOMEZ",              "MESERA",                   "AHORROS DAVIVIENDA", "4884-4926-8538", "SAVIA SALUD", 1750905, "empleado"),
    "KAROL QUINTERO":       ("1045018453", "KAROL DAYANA QUINTERO AGUDELO",      "MESERA",                   "AHORROS DAVIVIENDA", "3974-0008-5007", "SAVIA SALUD", 1750905, "empleado"),
    "KATERINE SEPULVEDA":   ("1000211127", "KATERIN MARYORY SEPULVEDA CRUZ",     "MESERA",                   "AHORROS DAVIVIENDA", "4884-4926-8785", "SAVIA SALUD", 1750905, "empleado"),
    "MARIA LIGELLA LOTERO": ("43844703",   "MARIA LIGELLA LOTERO",               "AUXILIAR ADMINISTRATIVA",  "AHORROS DAVIVIENDA", "4884-5662-0761", "SAVIA SALUD", 1750905, "empleado"),
    "JULIANA GOMEZ":        ("",           "JULIANA GOMEZ",                      "MESERA",                   "AHORROS DAVIVIENDA", "",               "",            10000,    "prestador"),
    # Valentina: empleada que no registra en el reloj (administrativa)
    # Su pago se calcula manualmente y se agrega al resumen
    "VALENTINA GRANDA":     ("1000397698", "VALENTINA GRANDA AGUDELO",           "AUXILIAR ADMINISTRATIVA",  "AHORROS DAVIVIENDA", "4884-5408-1842", "SAVIA SALUD", 1750905, "empleado"),
}

EXCEL_BASE = datetime(1899, 12, 30)

# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def parse_date(s):
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M:%S"]:
        try: return datetime.strptime(str(s).strip(), fmt)
        except: pass
    raise ValueError(f"Fecha no reconocida: {s}")

def dt_frac(dt):
    "datetime → fracción de día Excel (0.75 = 18:00)"
    return (dt.hour * 3600 + dt.minute * 60 + dt.second) / 86400

def frac_to_hm(f):
    "0.75 → '18:00'"
    h = int(f * 24); m = round((f * 24 - h) * 60)
    return f"{h:02d}:{m:02d}"

def date_serial(dt):
    return (dt - EXCEL_BASE).days

def thin(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_bottom(color="888888"):
    b = Side(style="medium", color=color)
    n = Side(style="none")
    return Border(bottom=b, left=n, right=n, top=n)

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="000000", size=10, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sc(ws, r, c, v, bold=False, bg=None, fc="000000", ha="left",
       fmt=None, italic=False, b=True, size=10, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = fnt(bold=bold, color=fc, size=size, italic=italic)
    cell.alignment = aln(h=ha, v="center", wrap=wrap)
    if bg: cell.fill = fill(bg)
    if b:  cell.border = thin()
    if fmt: cell.number_format = fmt
    return cell

C = dict(
    h1="1B3A5C", h2="2E6DA4", h_verde="1E8449", h_rojo="922B21",
    alt="EEF4FB", verde_c="D5F5E3", rojo_c="FADBD8", naranja="E59866",
    amarillo="FFF3CD", azul_c="D6EAF8", blanco="FFFFFF", gris="BDC3C7",
    # colilla
    col_h="1A237E",     # azul muy oscuro header empresa
    col_sub="283593",
    col_linea="3949AB",
    col_dev="E8F5E9",   # devengado verde suave
    col_ded="FFEBEE",   # deducciones rojo suave
    col_total="E3F2FD", # totales azul suave
    col_neto="1B3A5C",  # neto a pagar header
)

# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESAMIENTO MARCACIONES
# ═══════════════════════════════════════════════════════════════════════════════

def procesar(df, p_ini, p_fin):
    by_emp = {}
    for _, row in df.iterrows():
        try: fecha = parse_date(str(row['Tiempo']))
        except: continue
        num = str(row.get('Número', row.get('Numero', ''))).strip()
        nom = str(row.get('Nombre', '')).strip()
        est = str(row.get('Estado', '')).strip()
        k = f"{num}_{nom}"
        if k not in by_emp:
            by_emp[k] = {'id': num, 'nombre': nom, 'marcaciones': []}
        by_emp[k]['marcaciones'].append({'fecha': fecha, 'estado': est})

    resultados = []
    for k, emp in by_emp.items():
        dias_map = {}
        for m in emp['marcaciones']:
            dk = m['fecha'].strftime('%Y-%m-%d')
            if dk not in dias_map: dias_map[dk] = {'E': [], 'S': []}
            if m['estado'] == 'Entrada': dias_map[dk]['E'].append(m['fecha'])
            else: dias_map[dk]['S'].append(m['fecha'])

        dias = []
        cur = p_ini
        while cur <= p_fin:
            dk = cur.strftime('%Y-%m-%d')
            raw = dias_map.get(dk, {'E': [], 'S': []})
            entrada = salida = None
            ef = sf = 'ok'

            if raw['E']: entrada = sorted(raw['E'])[0]
            else: ef = 'missing'

            if raw['S']: salida = sorted(raw['S'])[-1]
            else:
                nk = (cur + timedelta(1)).strftime('%Y-%m-%d')
                early = [s for s in dias_map.get(nk, {'S': []})['S'] if s.hour < HORA_MADRUGADA]
                if early: salida = sorted(early)[-1]
                else: sf = 'missing'

            trab = noct = 0.0
            if entrada and salida and salida > entrada:
                trab = (salida - entrada).total_seconds() / 86400
                ns = entrada.replace(hour=HORA_NOCT_INI, minute=0, second=0)
                if salida > ns:
                    noct = (salida - max(entrada, ns)).total_seconds() / 86400

            tiene = bool(raw['E'] or raw['S'])
            dias.append({'fecha': cur, 'dk': dk, 'entrada': entrada, 'salida': salida,
                         'ef': ef, 'sf': sf, 'trab': trab, 'noct': noct, 'tiene': tiene})
            cur += timedelta(1)

        resultados.append({**emp, 'dias': dias})

    return resultados


def calcular(emp, salario, tipo="empleado"):
    """
    tipo='empleado'  → contrato laboral, salario/220 = valor hora, deducciones salud/pension
    tipo='prestador' → salario = valor_hora fijo (ej $10,000/h), sin deducciones
    """
    QDIA = HORAS_QUINCENA / 24
    MAXE = MAX_EXTRAS_NOM / 24

    if tipo == "prestador":
        # Para prestadores: salario ES el valor por hora directamente
        vh = salario
    else:
        vh = salario / HORAS_MES

    tot  = sum(d['trab'] for d in emp['dias'])
    noct = sum(d['noct'] for d in emp['dias'])

    if tipo == "prestador":
        # Prestador: pago por horas trabajadas, sin comparar vs 88h
        ext  = 0.0
        deu  = 0.0
        en   = 0.0
        ee   = 0.0
        # valor total = horas trabajadas × valor_hora
        val_total_prest = tot * 24 * vh
    else:
        ext  = max(0.0, tot - QDIA)
        deu  = max(0.0, QDIA - tot)
        en   = min(ext, MAXE)
        ee   = max(0.0, ext - MAXE)
        val_total_prest = 0.0

    dias_trab = sum(1 for d in emp['dias'] if d['trab'] > 0)

    return dict(salario=salario, vh=vh, tipo=tipo,
                tot=tot, noct=noct, ext=ext, deu=deu,
                en=en, ee=ee,
                en_h=en*24, ee_h=ee*24, noct_h=noct*24,
                deu_h=deu*24, tot_h=tot*24,
                val_en=en*24*vh*FACTOR_EXTRA,
                val_ee=ee*24*vh*FACTOR_EXTRA,
                val_noct=noct*24*vh*FACTOR_NOCT,
                val_deu=deu*24*vh,
                val_total_prest=val_total_prest,
                dias_trab=dias_trab)


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORTE HORARIOS
# ═══════════════════════════════════════════════════════════════════════════════

def hoja_empleado_horario(wb, emp, t, p_ini, p_fin):
    nom = emp['nombre']
    ws = wb.create_sheet(title=nom.split()[-1][:10].upper())
    ws.sheet_view.showGridLines = False
    DIAS = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]

    # Título
    ws.merge_cells("A1:K1")
    c = ws["A1"]; c.value = f"REPORTE DE HORARIOS — {nom.upper()}"
    c.font = fnt(bold=True, color="FFFFFF", size=12)
    c.fill = fill(C['h1']); c.alignment = aln("center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:C2"); ws["A2"].value = "Nombre"
    ws["A2"].font = fnt(bold=True); ws["A2"].fill = fill(C['alt'])
    ws.merge_cells("D2:K2"); ws["D2"].value = nom
    ws["D2"].font = fnt(bold=True, color=C['h2']); ws["D2"].fill = fill(C['alt'])

    hdrs = ["ID","Fecha","Turno","Entrada","Salida","Red. Ent.","Red. Sal.","Trabajado","T/T","Descanso","RECARGO\nNOCT."]
    for col, h in enumerate(hdrs, 1):
        c = sc(ws, 3, col, h, bold=True, bg=C['h2'], fc="FFFFFF", ha="center")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    fila = 4
    for i, d in enumerate(emp['dias']):
        bg = C['alt'] if i % 2 == 0 else C['blanco']
        if d['tiene'] and (d['ef'] == 'missing' or d['sf'] == 'missing'):
            bg = C['naranja']

        eid = int(emp['id']) if emp['id'].isdigit() else emp['id']
        sc(ws, fila, 1, eid, bg=bg, ha="center")
        sc(ws, fila, 2, d['fecha'], bg=bg, ha="center", fmt="DD/MM/YYYY")
        sc(ws, fila, 3, "HORARIO FLEXIBLE(00:00-23:59)", bg=bg)

        if not d['tiene']:
            sc(ws, fila, 4, "DESCANSO", bg=C['amarillo'], ha="center", bold=True)
            for col in range(5, 12): sc(ws, fila, col, "", bg=bg)
        else:
            e_str = frac_to_hm(dt_frac(d['entrada'])) if d['ef']=='ok' else "00:00 ⚠"
            s_str = frac_to_hm(dt_frac(d['salida']))  if d['sf']=='ok' else "00:00 ⚠"
            sc(ws, fila, 4, e_str, bg=bg if d['ef']=='ok' else C['naranja'], ha="center",
               fc="000000" if d['ef']=='ok' else "FFFFFF", bold=(d['ef']!='ok'))
            sc(ws, fila, 5, s_str, bg=bg if d['sf']=='ok' else C['naranja'], ha="center",
               fc="000000" if d['sf']=='ok' else "FFFFFF", bold=(d['sf']!='ok'))
            for col in [6, 7, 9, 10]: sc(ws, fila, col, "", bg=bg)
            sc(ws, fila, 8, round(d['trab'], 8), bg=bg, ha="center", fmt="0.00000000")
            nv = round(d['noct'], 8) if d['noct'] > 0 else ""
            sc(ws, fila, 11, nv, bg=C['azul_c'] if d['noct']>0 else bg, ha="center",
               fmt="0.00000000" if d['noct']>0 else None)
        ws.row_dimensions[fila].height = 15
        fila += 1

    # Totales
    for col, v in enumerate(["","","TOTAL HORAS LABORADAS","","","","",
                               round(t['tot'],8),"RECARGO NOCTURNO","",round(t['noct'],8)], 1):
        c = sc(ws, fila, col, v, bold=True, bg=C['h1'], fc="FFFFFF",
               ha="center" if col in [4,8,11] else "left",
               fmt="0.00000000" if col in [8,11] else None)
    ws.row_dimensions[fila].height = 18; fila += 1

    for col, v in enumerate(["","","HORAS QUE DEBE","","","","",
                               round(t['deu'],8) if t['deu']>0.0001 else 0,
                               "RECARGO QUE DEBE","",0], 1):
        sc(ws, fila, col, v, bold=(col in [3,9]), bg=C['rojo_c'],
           ha="center" if col in [8,11] else "left",
           fmt="0.00000000" if col in [8,11] else None)
    fila += 1

    for col, v in enumerate(["","","TOTAL","","","","",
                               round(t['tot'],8),"TOTAL","",round(t['noct'],8)], 1):
        sc(ws, fila, col, v, bold=(col in [3,9]), bg=C['verde_c'],
           ha="center" if col in [8,11] else "left",
           fmt="0.00000000" if col in [8,11] else None)
    fila += 2

    QDIA = HORAS_QUINCENA / 24
    # Bloque quincena/extras
    def mk(r, c1, l1, c2, v2, bgL=C['azul_c'], bgV=C['blanco'], fmt=None, boldV=False):
        sc(ws, r, c1, l1, bold=True, bg=bgL, b=True)
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c1+1)
        sc(ws, r, c2, v2, bg=bgV, ha="right", fmt=fmt, bold=boldV)

    mk(fila, 2, "HORAS POR QUINCENA", 4, "", bgV=C['azul_c'])
    sc(ws, fila, 5, "HORAS QUE DEBE", bold=True, bg=C['rojo_c'])
    ws.merge_cells(start_row=fila, start_column=5, end_row=fila, end_column=7)
    deuda_disp = -round(t['ext'],8) if t['ext']>0 else round(t['deu'],8)
    sc(ws, fila, 8, deuda_disp, bold=True,
       bg=C['rojo_c'] if t['deu']>0 else C['verde_c'], ha="right", fmt="0.00000000")
    fila += 1
    sc(ws, fila, 2, round(QDIA, 8), bold=True, bg=C['azul_c'], ha="right", fmt="0.00000000")
    sc(ws, fila, 5, "HORAS EXTRAS", bold=True, bg=C['verde_c'])
    ws.merge_cells(start_row=fila, start_column=5, end_row=fila, end_column=7)
    sc(ws, fila, 8, round(t['ext'],8), bold=True, bg=C['verde_c'], ha="right", fmt="0.00000000")
    fila += 1
    sc(ws, fila, 2, "SUELDO", bold=True, bg=C['azul_c'])
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
    fila += 1
    sc(ws, fila, 2, t['salario'], bold=True, ha="right", fmt="$#,##0")
    fila += 2

    # Tabla horas extras
    for col, h in enumerate(["","HORAS EXTRAS","","TIEMPO (días)","VALOR"], 1):
        sc(ws, fila, col, h, bold=True, bg=C['h2'], fc="FFFFFF", ha="center")
    fila += 1
    for lbl, td, val, bgc in [
        ("SE PAGAN EN COLILLA DE PAGO", t['en'], t['val_en'], C['verde_c']),
        ("SE PAGA EN EFECTIVO",         t['ee'], t['val_ee'], C['amarillo']),
    ]:
        sc(ws, fila, 2, lbl, bg=bgc)
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        sc(ws, fila, 4, round(td,8) if td else 0, bg=bgc, ha="right", fmt="0.00000000")
        sc(ws, fila, 5, round(val) if val else 0, bg=bgc, ha="right", fmt="$#,##0")
        ws.row_dimensions[fila].height = 15; fila += 1
    fila += 1

    # Tabla conceptos
    for col, h in enumerate(["","CONCEPTO","","VALOR HORA","TIEMPO (h)","TOTAL"], 1):
        sc(ws, fila, col, h, bold=True, bg=C['h1'], fc="FFFFFF", ha="center")
    fila += 1
    for lbl, vh_f, th, total, bgc in [
        ("Hora extra diurna",  t['vh']*FACTOR_EXTRA, t['ext']*24, t['val_en']+t['val_ee'], C['alt']),
        ("Recargo nocturno",   t['vh']*FACTOR_NOCT,  t['noct_h'],  t['val_noct'],           C['azul_c']),
    ]:
        sc(ws, fila, 2, lbl, bg=bgc)
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        sc(ws, fila, 4, round(vh_f,10), bg=bgc, ha="right", fmt="0.0000000000")
        sc(ws, fila, 5, round(th,8),    bg=bgc, ha="right", fmt="0.00000000")
        sc(ws, fila, 6, round(total),   bg=bgc, ha="right", fmt="$#,##0")
        ws.row_dimensions[fila].height = 15; fila += 1
    sc(ws, fila, 5, "TOTAL", bold=True, bg=C['h1'], fc="FFFFFF", ha="right")
    sc(ws, fila, 6, round(t['val_en']+t['val_ee']+t['val_noct']), bold=True,
       bg=C['h1'], fc="FFFFFF", ha="right", fmt="$#,##0")
    ws.row_dimensions[fila].height = 18; fila += 2

    # Novedades
    for col, h in enumerate(["","NOVEDAD","","DÍAS","% PAGO","VALOR"], 1):
        sc(ws, fila, col, h, bold=True, bg=C['h2'], fc="FFFFFF", ha="center")
    fila += 1
    for nov in ["Incapacidad (0% / 66.66% / 100%)", "Licencia no remunerada",
                "Vacaciones", "Día de la familia / compensatorio"]:
        sc(ws, fila, 2, nov, bg=C['amarillo'])
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        sc(ws, fila, 4, 0,      bg=C['amarillo'], ha="center")
        sc(ws, fila, 5, "100%", bg=C['amarillo'], ha="center")
        sc(ws, fila, 6, 0,      bg=C['amarillo'], ha="right", fmt="$#,##0")
        ws.row_dimensions[fila].height = 15; fila += 1

    widths = [4, 8, 24, 10, 10, 11, 11, 14, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def crear_reporte_horarios(resultados_t, p_ini, p_fin, nombre_out):
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ── Hoja resumen ──
    ws = wb.create_sheet("RESUMEN NOMINA", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L1")
    c = ws["A1"]; c.value = f"RESUMEN DE NÓMINA — {p_ini.strftime('%d/%m/%Y')} AL {p_fin.strftime('%d/%m/%Y')}"
    c.font = fnt(bold=True, color="FFFFFF", size=13); c.fill = fill(C['h1']); c.alignment = aln("center")
    ws.row_dimensions[1].height = 28

    hdrs = ["Nombre","Salario\nMensual","Total H.\nTrabajadas","Horas\nExtras",
            "Ext.\nNómina","Ext.\nEfectivo","H. Recargo\nNocturno","H.\nDebe",
            "Valor Ext.\nNómina","Valor Ext.\nEfectivo","Valor\nRecargo","TOTAL\nQUINCENA"]
    for col, h in enumerate(hdrs, 1):
        c = sc(ws, 2, col, h, bold=True, bg=C['h2'], fc="FFFFFF", ha="center")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    for i, (emp, t) in enumerate(resultados_t):
        r = i + 3; bg = C['alt'] if i % 2 == 0 else C['blanco']
        total_q = (t['salario']/2) + t['val_en'] + t['val_noct'] - t['val_deu']
        vals = [emp['nombre'], t['salario'], round(t['tot_h'],2), round(t['ext']*24,2),
                round(t['en_h'],2), round(t['ee_h'],2), round(t['noct_h'],2), round(t['deu_h'],2),
                round(t['val_en']), round(t['val_ee']), round(t['val_noct']), round(total_q)]
        fmts = [None,'$#,##0','0.00','0.00','0.00','0.00','0.00','0.00',
                '$#,##0','$#,##0','$#,##0','$#,##0']
        for col, (v, fm) in enumerate(zip(vals, fmts), 1):
            cbg = bg
            if col == 4 and t['ext'] > 0: cbg = C['verde_c']
            if col == 8 and t['deu_h'] > 0.1: cbg = C['rojo_c']
            sc(ws, r, col, v, bg=cbg, ha="right" if col > 1 else "left", fmt=fm)
        ws.row_dimensions[r].height = 15

    n = len(resultados_t); tr = n + 3
    sc(ws, tr, 1, "TOTALES", bold=True, bg=C['h1'], fc="FFFFFF")
    for col in range(2, 13):
        lt = get_column_letter(col)
        fm = '$#,##0' if col in [2,9,10,11,12] else '0.00'
        c = ws.cell(tr, col, value=f"=SUM({lt}3:{lt}{tr-1})")
        c.font = fnt(bold=True, color="FFFFFF"); c.fill = fill(C['h1'])
        c.alignment = aln("right"); c.border = thin(); c.number_format = fm
    ws.row_dimensions[tr].height = 18
    for i, w in enumerate([26,13,12,10,10,10,13,10,14,14,13,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # ── Hoja extras/recargos ──
    ws2 = wb.create_sheet("HORAS EXTRAS Y RECARGOS")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:I1")
    c = ws2["A1"]; c.value = "RESUMEN HORAS EXTRAS, RECARGO NOCTURNO Y TIEMPO ADEUDADO"
    c.font = fnt(bold=True, color="FFFFFF", size=12); c.fill = fill(C['h1']); c.alignment = aln("center")
    ws2.row_dimensions[1].height = 24
    hdrs2 = ["Nombre","H. Extras\nNómina","Valor\nExtras Nom.","H. Extras\nEfectivo",
             "Valor\nExtras Ef.","H. Recargo\nNocturno","Valor\nRecargo Noct.","H. Adeudadas","Valor\nAdeudo"]
    for col, h in enumerate(hdrs2, 1):
        c = sc(ws2, 2, col, h, bold=True, bg=C['h2'], fc="FFFFFF", ha="center")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[2].height = 30
    for i, (emp, t) in enumerate(resultados_t):
        r = i + 3; bg = C['alt'] if i % 2 == 0 else C['blanco']
        vals = [emp['nombre'], round(t['en_h'],2), round(t['val_en']), round(t['ee_h'],2),
                round(t['val_ee']), round(t['noct_h'],2), round(t['val_noct']),
                round(t['deu_h'],2), round(t['val_deu'])]
        fmts = [None,'0.00','$#,##0','0.00','$#,##0','0.00','$#,##0','0.00','$#,##0']
        for col, (v, fm) in enumerate(zip(vals, fmts), 1):
            cbg = bg
            if col == 2 and t['en_h'] > 0: cbg = C['verde_c']
            if col == 4 and t['ee_h'] > 0: cbg = C['amarillo']
            if col == 8 and t['deu_h'] > 0.1: cbg = C['rojo_c']
            sc(ws2, r, col, v, bg=cbg, ha="right" if col > 1 else "left", fmt=fm)
        ws2.row_dimensions[r].height = 15
    n2 = len(resultados_t); tr2 = n2 + 3
    sc(ws2, tr2, 1, "TOTALES", bold=True, bg=C['h1'], fc="FFFFFF")
    for col in range(2, 10):
        lt = get_column_letter(col)
        fm = '$#,##0' if col in [3,5,7,9] else '0.00'
        c = ws2.cell(tr2, col, value=f"=SUM({lt}3:{lt}{tr2-1})")
        c.font = fnt(bold=True, color="FFFFFF"); c.fill = fill(C['h1'])
        c.alignment = aln("right"); c.border = thin(); c.number_format = fm
    for i, w in enumerate([26,11,15,12,15,13,16,11,14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    # ── Hojas individuales ──
    for emp, t in resultados_t:
        hoja_empleado_horario(wb, emp, t, p_ini, p_fin)

    wb.save(nombre_out)
    print(f"✅ {nombre_out}")


# ═══════════════════════════════════════════════════════════════════════════════
#  COLILLA DE PAGO
# ═══════════════════════════════════════════════════════════════════════════════

def colilla_empleado(wb, emp, t, p_ini, p_fin, datos_col):
    cedula, nombre_completo, cargo, banco, cuenta, eps, _, tipo_col = datos_col
    sal = t['salario']
    AUX_Q = AUXILIO_TRANSPORTE_ANUAL / 2
    dias_trab = t['dias_trab']

    if tipo_col == "prestador":
        # Prestador: pago por horas, sin auxilio transporte ni deducciones
        sal_q     = 0
        aux_transp = 0
        dev_ext   = 0
        dev_noct  = 0
        pension   = 0
        salud     = 0
        total_dev = t['val_total_prest']
        total_ded = 0
        neto      = total_dev
        # Para mostrar en la colilla el concepto correcto
        cant_prest = round(t['tot_h'], 10)
    else:
        aux_transp = AUX_Q * dias_trab / 15
        sal_q     = sal / 2
        dev_ext   = t['val_en']
        dev_noct  = t['val_noct']
        ibc       = sal_q + dev_ext + dev_noct
        pension   = ibc * PENSION_PCT
        salud     = ibc * SALUD_PCT
        total_dev = sal_q + aux_transp + dev_ext + dev_noct
        total_ded = pension + salud
        neto      = total_dev - total_ded
        cant_prest = 0

    # Fechas
    serial_hoy = date_serial(datetime.now())
    serial_ini = date_serial(p_ini)
    serial_fin = date_serial(p_fin)
    mes = p_ini.month
    anio = p_ini.year

    nom_hoja = nombre_completo.split()[0][:10].upper()
    # Avoid duplicate sheet names
    base = nom_hoja; i = 1
    while base in [ws.title for ws in wb.worksheets]:
        base = f"{nom_hoja}{i}"; i += 1
    ws = wb.create_sheet(title=base)
    ws.sheet_view.showGridLines = False

    # ── Colilla se imprime DOS VECES en la misma hoja ──
    for bloque in range(2):
        R = bloque * 22 + 1  # fila de inicio de cada bloque

        def line(row, col, v, bold=False, bg=None, fc="000000", ha="left",
                 fmt=None, merge_end=None, b=True, size=9):
            r = R + row
            c = sc(ws, r, col, v, bold=bold, bg=bg, fc=fc, ha=ha,
                   fmt=fmt, b=b, size=size)
            if merge_end:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=merge_end)
            return c

        # ── Header empresa ──
        ws.row_dimensions[R].height = 20
        line(0, 1, f"{EMPRESA['nombre']}  NIT: {EMPRESA['nit']}  DIR: {EMPRESA['dir']}  TEL: {EMPRESA['tel']}",
             bold=True, bg=C['col_h'], fc="FFFFFF", ha="left", merge_end=11, size=9)
        ws.cell(R, 12).value = serial_hoy
        ws.cell(R, 12).number_format = "DD/MM/YYYY"
        ws.cell(R, 12).font = fnt(bold=True, color="FFFFFF", size=9)
        ws.cell(R, 12).fill = fill(C['col_h'])
        ws.cell(R, 12).alignment = aln("right")
        ws.cell(R, 12).border = thin()

        # ── Fila periodo ──
        ws.row_dimensions[R+1].height = 16
        line(1, 1, "Periodo", bold=True, bg=C['col_sub'], fc="FFFFFF", size=9)
        ws.cell(R+1, 2).value = mes
        ws.cell(R+1, 2).font = fnt(bold=True, color="FFFFFF", size=9)
        ws.cell(R+1, 2).fill = fill(C['col_sub']); ws.cell(R+1, 2).border = thin()
        ws.cell(R+1, 3).value = anio
        ws.cell(R+1, 3).font = fnt(bold=True, color="FFFFFF", size=9)
        ws.cell(R+1, 3).fill = fill(C['col_sub']); ws.cell(R+1, 3).border = thin()

        line(1, 4, "Nómina Entre", bold=True, bg=C['col_sub'], fc="FFFFFF", size=9)
        ws.cell(R+1, 5).value = serial_ini
        ws.cell(R+1, 5).number_format = "DD/MM/YYYY"
        ws.cell(R+1, 5).font = fnt(color="FFFFFF", size=9)
        ws.cell(R+1, 5).fill = fill(C['col_sub']); ws.cell(R+1, 5).border = thin()
        line(1, 6, "y", bg=C['col_sub'], fc="FFFFFF", ha="center", size=9)
        ws.cell(R+1, 7).value = serial_fin
        ws.cell(R+1, 7).number_format = "DD/MM/YYYY"
        ws.cell(R+1, 7).font = fnt(color="FFFFFF", size=9)
        ws.cell(R+1, 7).fill = fill(C['col_sub']); ws.cell(R+1, 7).border = thin()
        for col in range(8, 13):
            ws.cell(R+1, col).fill = fill(C['col_sub']); ws.cell(R+1, col).border = thin()

        # ── Empleado / Cargo / Salario ──
        ws.row_dimensions[R+2].height = 16
        line(2, 1, "EMPLEADO:", bold=True, size=9)
        ws.cell(R+2, 2).value = cedula
        ws.cell(R+2, 2).font = fnt(size=9); ws.cell(R+2, 2).border = thin()
        ws.merge_cells(start_row=R+2, start_column=3, end_row=R+2, end_column=8)
        ws.cell(R+2, 3).value = nombre_completo.upper()
        ws.cell(R+2, 3).font = fnt(bold=True, size=9); ws.cell(R+2, 3).border = thin()
        ws.cell(R+2, 3).alignment = aln("center")
        ws.cell(R+2, 9).value = "SALARIO MENSUAL:"
        ws.cell(R+2, 9).font = fnt(bold=True, size=9); ws.cell(R+2, 9).border = thin()
        ws.merge_cells(start_row=R+2, start_column=9, end_row=R+2, end_column=10)
        ws.cell(R+2, 11).value = sal
        ws.cell(R+2, 11).font = fnt(bold=True, size=9); ws.cell(R+2, 11).border = thin()
        ws.cell(R+2, 11).number_format = "$#,##0"; ws.cell(R+2, 11).alignment = aln("right")
        ws.cell(R+2, 12).border = thin()

        ws.row_dimensions[R+3].height = 16
        line(3, 1, "CARGO:", bold=True, size=9)
        ws.merge_cells(start_row=R+3, start_column=2, end_row=R+3, end_column=6)
        ws.cell(R+3, 2).value = cargo
        ws.cell(R+3, 2).font = fnt(size=9); ws.cell(R+3, 2).border = thin()
        for col in range(7, 13): ws.cell(R+3, col).border = thin()

        # ── Separador ──
        ws.row_dimensions[R+4].height = 6
        ws.merge_cells(start_row=R+4, start_column=1, end_row=R+4, end_column=12)
        ws.cell(R+4, 1).fill = fill(C['col_h']); ws.cell(R+4, 1).border = thin()

        # ── Cabecera tabla ──
        ws.row_dimensions[R+5].height = 18
        for col, h in enumerate(["CODIGO","","","DESCRIPCION","","","DOC","CANT","DEVENGADO","DEDUCCION","","SALDO"], 1):
            c = sc(ws, R+5, col, h, bold=True, bg=C['col_linea'], fc="FFFFFF", ha="center", size=9)
        ws.merge_cells(start_row=R+5, start_column=1, end_row=R+5, end_column=3)
        ws.merge_cells(start_row=R+5, start_column=4, end_row=R+5, end_column=6)
        ws.merge_cells(start_row=R+5, start_column=10, end_row=R+5, end_column=11)

        # ── Filas de conceptos ──
        if tipo_col == "prestador":
            conceptos = [
                (1,  "HORAS TRABAJADAS",               round(t['tot_h'],10), t['val_total_prest'], None, C['col_dev']),
                (2,  "AUXILIO DE TRANSPORTE",           0,              0,         None,    C['col_dev']),
                (3,  "AUXILIO  ",                       0,              0,         None,    C['col_dev']),
                (4,  "RECARGO NOCTURNO",                0,              0,         None,    C['col_dev']),
                (121,"DEDUCCION DE PRESTAMOS",          None,           None,      0,       C['col_ded']),
                (123,"DEDUCCION PENSION",               None,           None,      0,       C['col_ded']),
                (127,"DEDUCCION SALUD",                 None,           None,      0,       C['col_ded']),
            ]
        else:
            conceptos = [
                (1,  "SALARIO BASICO",                 110,             sal_q,     None,    C['col_dev']),
                (2,  "AUXILIO DE TRANSPORTE",           dias_trab,      aux_transp, None,   C['col_dev']),
                (3,  "AUXILIO  ",                       round(t['en_h'],10), dev_ext, None, C['col_dev']),
                (4,  "RECARGO NOCTURNO",               round(t['noct_h'],10), dev_noct, None, C['col_dev']),
                (121,"DEDUCCION DE PRESTAMOS",          None,           None,      0,       C['col_ded']),
                (123,f"DEDUCCION PENSION PROTECCIÓN 4%", None,          None,      pension, C['col_ded']),
                (127,f"DEDUCCION SALUD 4% {eps.upper()}", None,         None,      salud,   C['col_ded']),
            ]

        for frow, (cod, desc, cant, dev, ded, bgc) in enumerate(conceptos):
            r = R + 6 + frow
            ws.row_dimensions[r].height = 15
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r,1).value = cod; ws.cell(r,1).font = fnt(bold=True, size=9)
            ws.cell(r,1).fill = fill(bgc); ws.cell(r,1).border = thin()
            ws.cell(r,1).alignment = aln("center")
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
            ws.cell(r,4).value = desc; ws.cell(r,4).font = fnt(size=9)
            ws.cell(r,4).fill = fill(bgc); ws.cell(r,4).border = thin()
            ws.cell(r,7).border = thin(); ws.cell(r,7).fill = fill(bgc)  # DOC
            ws.cell(r,8).value = cant if cant else ""
            ws.cell(r,8).font = fnt(size=9); ws.cell(r,8).fill = fill(bgc)
            ws.cell(r,8).border = thin(); ws.cell(r,8).alignment = aln("right")
            ws.cell(r,8).number_format = "0.##########"
            ws.cell(r,9).value = dev if dev else ""
            ws.cell(r,9).font = fnt(size=9); ws.cell(r,9).fill = fill(bgc)
            ws.cell(r,9).border = thin(); ws.cell(r,9).alignment = aln("right")
            ws.cell(r,9).number_format = "$#,##0.##"
            ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
            ws.cell(r,10).value = ded if ded else ""
            ws.cell(r,10).font = fnt(size=9); ws.cell(r,10).fill = fill(bgc)
            ws.cell(r,10).border = thin(); ws.cell(r,10).alignment = aln("right")
            ws.cell(r,10).number_format = "$#,##0.##"
            ws.cell(r,12).border = thin(); ws.cell(r,12).fill = fill(bgc)

        # ── Fila TOTALES ──
        rt = R + 13; ws.row_dimensions[rt].height = 17
        ws.merge_cells(start_row=rt, start_column=1, end_row=rt, end_column=6)
        ws.cell(rt,1).value = "TOTALES"
        ws.cell(rt,1).font = fnt(bold=True, size=9, color="FFFFFF")
        ws.cell(rt,1).fill = fill(C['col_neto']); ws.cell(rt,1).border = thin()
        ws.cell(rt,1).alignment = aln("right")
        ws.cell(rt,7).border = thin(); ws.cell(rt,7).fill = fill(C['col_total'])
        ws.cell(rt,8).border = thin(); ws.cell(rt,8).fill = fill(C['col_total'])
        ws.cell(rt,9).value = total_dev
        ws.cell(rt,9).font = fnt(bold=True, size=9); ws.cell(rt,9).fill = fill(C['col_total'])
        ws.cell(rt,9).border = thin(); ws.cell(rt,9).alignment = aln("right")
        ws.cell(rt,9).number_format = "$#,##0.##"
        ws.merge_cells(start_row=rt, start_column=10, end_row=rt, end_column=11)
        ws.cell(rt,10).value = total_ded
        ws.cell(rt,10).font = fnt(bold=True, size=9); ws.cell(rt,10).fill = fill(C['col_total'])
        ws.cell(rt,10).border = thin(); ws.cell(rt,10).alignment = aln("right")
        ws.cell(rt,10).number_format = "$#,##0.##"
        ws.cell(rt,12).border = thin(); ws.cell(rt,12).fill = fill(C['col_total'])

        # ── NETO A PAGAR ──
        rn = R + 14; ws.row_dimensions[rn].height = 20
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=8)
        ws.cell(rn,1).value = "NETO A PAGAR               "
        ws.cell(rn,1).font = fnt(bold=True, size=10, color="FFFFFF")
        ws.cell(rn,1).fill = fill(C['col_neto']); ws.cell(rn,1).border = thin()
        ws.cell(rn,1).alignment = aln("right")
        ws.merge_cells(start_row=rn, start_column=9, end_row=rn, end_column=12)
        ws.cell(rn,9).value = neto
        ws.cell(rn,9).font = fnt(bold=True, size=11, color="FFFFFF")
        ws.cell(rn,9).fill = fill(C['col_neto']); ws.cell(rn,9).border = thin()
        ws.cell(rn,9).alignment = aln("right")
        ws.cell(rn,9).number_format = "$#,##0.##"

        # ── Firma ──
        rf = R + 15; ws.row_dimensions[rf].height = 15
        ws.cell(rf,1).value = nombre_completo.upper()
        ws.cell(rf,1).font = fnt(bold=True, size=9); ws.cell(rf,1).border = thin()
        ws.merge_cells(start_row=rf, start_column=1, end_row=rf, end_column=6)

        rc = R + 16; ws.row_dimensions[rc].height = 15
        ws.cell(rc,1).value = f"C.C  "; ws.cell(rc,1).font = fnt(bold=True, size=9)
        ws.cell(rc,1).border = thin()
        ws.cell(rc,2).value = cedula; ws.cell(rc,2).font = fnt(size=9); ws.cell(rc,2).border = thin()

        rb = R + 17; ws.row_dimensions[rb].height = 15
        ws.cell(rb,1).value = "Cuenta:   "; ws.cell(rb,1).font = fnt(bold=True, size=9)
        ws.cell(rb,1).border = thin()
        ws.cell(rb,2).value = banco; ws.cell(rb,2).font = fnt(size=9); ws.cell(rb,2).border = thin()
        ws.merge_cells(start_row=rb, start_column=2, end_row=rb, end_column=5)
        ws.cell(rb,6).value = cuenta; ws.cell(rb,6).font = fnt(size=9); ws.cell(rb,6).border = thin()
        ws.merge_cells(start_row=rb, start_column=6, end_row=rb, end_column=8)

        # Fila espaciadora entre bloques
        if bloque == 0:
            ws.row_dimensions[R+18].height = 10
            ws.row_dimensions[R+19].height = 4
            for col in range(1, 13):
                ws.cell(R+19, col).fill = fill(C['col_h']); ws.cell(R+19, col).border = thin()

    # Anchos de columna
    widths = [7, 9, 5, 12, 8, 8, 6, 12, 16, 14, 4, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Área de impresión
    ws.print_area = f"A1:L{R+17}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def neto_empleado(t, datos):
    """Calcula neto a pagar para un empleado (con deducciones)."""
    sal_q    = t['salario'] / 2
    dias     = t['dias_trab']
    aux_t    = (AUXILIO_TRANSPORTE_ANUAL / 2) * dias / 15
    ibc      = sal_q + t['val_en'] + t['val_noct']
    ded      = ibc * (PENSION_PCT + SALUD_PCT)
    total_dev = sal_q + aux_t + t['val_en'] + t['val_noct']
    return total_dev - ded


def crear_resumen_nomina(resultados_t, valentina_neto, p_ini, p_fin, nombre_out):
    """
    Genera RESUMEN_NOMINA con 4 secciones exactas del original:
    1. Nómina para consignar (empleados con contrato)
    2. Nómina en efectivo
    3. Horas extras pagadas en efectivo
    4. Prestadores de servicio
    """
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Hoja1")
    ws.sheet_view.showGridLines = False

    # Columnas
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16

    mes_str = p_ini.strftime('%B').upper()
    anio    = p_ini.year
    q_str   = f"1 al 15" if p_ini.day == 1 else f"16 al {p_fin.day}"
    periodo_label = f"TOTAL NOMINA {q_str} de {mes_str} de {anio}"

    def seccion_header(row, titulo):
        ws.merge_cells(f"B{row}:F{row}")
        c = ws.cell(row, 2, value=titulo)
        c.font = fnt(bold=True, color="FFFFFF", size=10)
        c.fill = fill(C['col_neto']); c.alignment = aln("center"); c.border = thin()
        ws.row_dimensions[row].height = 20

    def col_header(row):
        for col, h in enumerate(["CEDULA","NOMBRE","TIPO CUENTA","# CUENTA","VALOR A PAGAR"], 2):
            sc(ws, row, col, h, bold=True, bg=C['col_linea'], fc="FFFFFF", ha="center", size=9)
        ws.row_dimensions[row].height = 16

    def data_row(row, cedula, nombre, tipo_cta, cuenta, valor, bg):
        sc(ws, row, 2, cedula,   bg=bg, size=9)
        sc(ws, row, 3, nombre,   bg=bg, size=9)
        sc(ws, row, 4, tipo_cta, bg=bg, size=9)
        sc(ws, row, 5, cuenta,   bg=bg, size=9)
        sc(ws, row, 6, round(valor), bg=bg, ha="right", fmt="$#,##0", size=9)
        ws.row_dimensions[row].height = 15

    def total_row(row, valor):
        ws.merge_cells(f"B{row}:E{row}")
        ws.cell(row, 2).value = "TOTAL"
        ws.cell(row, 2).font = fnt(bold=True, color="FFFFFF", size=9)
        ws.cell(row, 2).fill = fill(C['h1']); ws.cell(row, 2).border = thin()
        ws.cell(row, 2).alignment = aln("right")
        sc(ws, row, 6, round(valor), bold=True, bg=C['h1'], fc="FFFFFF",
           ha="right", fmt="$#,##0", size=9)
        ws.row_dimensions[row].height = 16

    def spacer(row):
        ws.row_dimensions[row].height = 8

    r = 1

    # Título principal
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(r, 2, value="RESUMEN NOMINA PARA CONSIGNAR CON AUMENTOS")
    c.font = fnt(bold=True, size=11); c.alignment = aln("center")
    ws.row_dimensions[r].height = 22; r += 1

    # ── SECCIÓN 1: Nómina para consignar ──────────────────────────────────────
    col_header(r); r += 1

    total_consignar = 0
    empleados_consignar = []
    for emp, t in resultados_t:
        datos = COLABORADORES.get(emp['nombre'])
        if not datos or datos[7] != "empleado": continue
        cedula, nombre_c, cargo, banco, cuenta, eps, _, _ = datos
        neto = neto_empleado(t, datos)
        empleados_consignar.append((cedula, nombre_c, banco, cuenta, neto))
        total_consignar += neto

    # Agregar Valentina (empleada sin reloj)
    vd = COLABORADORES.get("VALENTINA GRANDA")
    if vd and valentina_neto > 0:
        empleados_consignar.append((vd[0], vd[1], vd[3], vd[4], valentina_neto))
        total_consignar += valentina_neto

    for i, (ced, nom, banco, cta, neto) in enumerate(empleados_consignar):
        bg = C['alt'] if i % 2 == 0 else C['blanco']
        data_row(r, ced, nom, banco, cta, neto, bg); r += 1

    total_row(r, total_consignar); r += 1; spacer(r); r += 1

    # ── SECCIÓN 2: Nómina en efectivo ─────────────────────────────────────────
    seccion_header(r, "RESUMEN NOMINA PARA PAGO EN EFECTIVO"); r += 1
    col_header(r); r += 1
    for _ in range(2):
        for col in range(2, 7): sc(ws, r, col, "", bg=C['blanco'], size=9)
        ws.row_dimensions[r].height = 15; r += 1
    total_row(r, 0); r += 1; spacer(r); r += 1

    # ── SECCIÓN 3: Horas extras en efectivo ───────────────────────────────────
    seccion_header(r, "HORAS EXTRAS PAGAS EN EFECTIVO"); r += 1
    for col, h in enumerate(["CEDULA","NOMBRE","","# CUENTA","VALOR A PAGAR"], 2):
        sc(ws, r, col, h, bold=True, bg=C['col_linea'], fc="FFFFFF", ha="center", size=9)
    ws.row_dimensions[r].height = 16; r += 1

    total_ef = 0
    filas_ef = []
    for emp, t in resultados_t:
        if t['val_ee'] > 0.5:
            datos = COLABORADORES.get(emp['nombre'])
            ced = datos[0] if datos else emp['id']
            nom = datos[1] if datos else emp['nombre']
            filas_ef.append((ced, nom, round(t['val_ee'])))
            total_ef += t['val_ee']

    if filas_ef:
        for i, (ced, nom, val) in enumerate(filas_ef):
            bg = C['amarillo'] if i % 2 == 0 else C['blanco']
            sc(ws, r, 2, ced,  bg=bg, size=9)
            sc(ws, r, 3, nom,  bg=bg, size=9)
            sc(ws, r, 4, "",   bg=bg, size=9)
            sc(ws, r, 5, "",   bg=bg, size=9)
            sc(ws, r, 6, val,  bg=bg, ha="right", fmt="$#,##0", size=9)
            ws.row_dimensions[r].height = 15; r += 1
    else:
        for col in range(2, 7): sc(ws, r, col, "", bg=C['blanco'], size=9)
        ws.row_dimensions[r].height = 15; r += 1

    total_row(r, total_ef); r += 1; spacer(r); r += 1

    # ── SECCIÓN 4: Prestadores de servicio ────────────────────────────────────
    seccion_header(r, "RESUMEN DE NOMINA PRESTADORES DE SERVICIO"); r += 1
    for col, h in enumerate(["CEDULA","NOMBRE","","# CUENTA","VALOR A PAGAR"], 2):
        sc(ws, r, col, h, bold=True, bg=C['col_linea'], fc="FFFFFF", ha="center", size=9)
    ws.row_dimensions[r].height = 16; r += 1

    total_prest = 0
    for emp, t in resultados_t:
        datos = COLABORADORES.get(emp['nombre'])
        if not datos or datos[7] != "prestador": continue
        ced = datos[0]; nom = datos[1]
        val = t['val_total_prest']
        total_prest += val
        sc(ws, r, 2, ced, bg=C['alt'],    size=9)
        sc(ws, r, 3, nom, bg=C['alt'],    size=9)
        sc(ws, r, 4, "",  bg=C['alt'],    size=9)
        sc(ws, r, 5, "",  bg=C['alt'],    size=9)
        sc(ws, r, 6, round(val), bg=C['alt'], ha="right", fmt="$#,##0", size=9)
        ws.row_dimensions[r].height = 15; r += 1

    # Fila vacía extra
    for col in range(2, 7): sc(ws, r, col, "", bg=C['blanco'], size=9)
    ws.row_dimensions[r].height = 15; r += 1
    total_row(r, total_prest); r += 1; spacer(r); r += 1

    # ── GRAN TOTAL ────────────────────────────────────────────────────────────
    gran_total = total_consignar + total_ef + total_prest
    ws.merge_cells(f"B{r}:E{r}")
    ws.cell(r, 2).value = periodo_label
    ws.cell(r, 2).font = fnt(bold=True, color="FFFFFF", size=10)
    ws.cell(r, 2).fill = fill(C['col_h']); ws.cell(r, 2).border = thin()
    ws.cell(r, 2).alignment = aln("right")
    sc(ws, r, 6, round(gran_total), bold=True, bg=C['col_h'], fc="FFFFFF",
       ha="right", fmt="$#,##0", size=10)
    ws.row_dimensions[r].height = 22

    ws.freeze_panes = "B2"
    wb.save(nombre_out)
    print(f"✅ {nombre_out}")
    return gran_total


def crear_colilla_pago(resultados_t, p_ini, p_fin, nombre_out):
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # Hoja resumen de pagos (para transferencias)
    ws = wb.create_sheet("RESUMEN PAGO", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = f"RESUMEN DE NÓMINA PARA PAGO — {p_ini.strftime('%d/%m/%Y')} AL {p_fin.strftime('%d/%m/%Y')}"
    c.font = fnt(bold=True, color="FFFFFF", size=12); c.fill = fill(C['col_neto']); c.alignment = aln("center")
    ws.row_dimensions[1].height = 26

    hdrs = ["Nombre Completo","Cédula","Banco","N° Cuenta","Total Devengado","Deducciones","NETO A PAGAR"]
    for col, h in enumerate(hdrs, 1):
        c = sc(ws, 2, col, h, bold=True, bg=C['col_linea'], fc="FFFFFF", ha="center")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    gran_neto = 0
    for i, (emp, t) in enumerate(resultados_t):
        r = i + 3; bg = C['alt'] if i % 2 == 0 else C['blanco']
        datos = COLABORADORES.get(emp['nombre'], (emp['id'], emp['nombre'], "", "", "", "SAVIA SALUD", t['salario'], "empleado"))
        cedula, nombre_c, cargo, banco, cuenta, eps, _, tipo_col = datos
        if tipo_col == "prestador":
            neto = t['val_total_prest']; ded = 0; total_dev = neto
        else:
            sal_q = t['salario'] / 2; dias_trab = t['dias_trab']
            aux_t = (AUXILIO_TRANSPORTE_ANUAL / 2) * dias_trab / 15
            ibc = sal_q + t['val_en'] + t['val_noct']
            ded = ibc * (PENSION_PCT + SALUD_PCT)
            total_dev = sal_q + aux_t + t['val_en'] + t['val_noct']
            neto = total_dev - ded
        gran_neto += neto

        vals = [nombre_c, cedula, banco, cuenta,
                round(total_dev), round(ded), round(neto)]
        fmts = [None, None, None, None, '$#,##0', '$#,##0', '$#,##0']
        for col, (v, fm) in enumerate(zip(vals, fmts), 1):
            sc(ws, r, col, v, bg=bg, ha="right" if col > 4 else "left", fmt=fm)
        ws.row_dimensions[r].height = 15

    # Totales
    n = len(resultados_t); tr = n + 3
    sc(ws, tr, 1, "TOTAL A TRANSFERIR", bold=True, bg=C['col_neto'], fc="FFFFFF")
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=6)
    sc(ws, tr, 7, round(gran_neto), bold=True, bg=C['col_neto'], fc="FFFFFF",
       ha="right", fmt='$#,##0')
    ws.row_dimensions[tr].height = 20

    for i, w in enumerate([30, 14, 20, 18, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    # Colillas individuales
    for emp, t in resultados_t:
        datos = COLABORADORES.get(emp['nombre'],
                   (emp['id'], emp['nombre'], "CARGO", "DAVIVIENDA", "", "SAVIA SALUD", t['salario'], "empleado"))
        colilla_empleado(wb, emp, t, p_ini, p_fin, datos)

    wb.save(nombre_out)
    print(f"✅ {nombre_out}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Nómina Soccer 7 — GRANDA VARGAS SAS")
    parser.add_argument("archivo", help="CSV del reloj biométrico Zkteco K50")
    parser.add_argument("--inicio", help="YYYY-MM-DD inicio período")
    parser.add_argument("--fin",    help="YYYY-MM-DD fin período")
    args = parser.parse_args()

    print(f"📂 Leyendo {args.archivo}...")
    df = None
    for sep in [',', ';', '\t']:
        try:
            tmp = pd.read_csv(args.archivo, sep=sep, encoding='utf-8-sig')
            if len(tmp.columns) >= 4: df = tmp; break
        except: pass
    if df is None: print("ERROR: No se pudo leer el CSV."); return

    fechas = sorted([parse_date(str(r)) for r in df['Tiempo']])
    if args.inicio:
        p_ini = datetime.strptime(args.inicio, "%Y-%m-%d")
    else:
        mf = fechas[0]
        p_ini = mf.replace(day=1 if mf.day <= 15 else 16, hour=0, minute=0, second=0)
    if args.fin:
        p_fin = datetime.strptime(args.fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    else:
        ult = calendar.monthrange(p_ini.year, p_ini.month)[1]
        p_fin = p_ini.replace(day=15 if p_ini.day == 1 else ult, hour=23, minute=59, second=59)

    print(f"   Período: {p_ini.strftime('%d/%m/%Y')} → {p_fin.strftime('%d/%m/%Y')}")
    print(f"   {df['Nombre'].nunique()} colaboradoras | {len(df)} marcaciones")

    resultados = procesar(df, p_ini, p_fin)
    resultados_t = []
    for emp in resultados:
        datos = COLABORADORES.get(emp['nombre'], (emp['id'], emp['nombre'], "", "", "", "SAVIA SALUD", SALARIO_MINIMO, "empleado"))
        sal  = datos[6]
        tipo = datos[7]
        t = calcular(emp, sal, tipo)
        resultados_t.append((emp, t))

    # Valentina: empleada sin reloj biométrico — calcular su neto fijo
    vd = COLABORADORES.get("VALENTINA GRANDA")
    valentina_neto = 0
    if vd:
        sal_v = vd[6]; sal_v_q = sal_v / 2
        aux_v = AUXILIO_TRANSPORTE_ANUAL / 2
        ibc_v = sal_v_q  # sin extras ni recargo para admin
        ded_v = ibc_v * (PENSION_PCT + SALUD_PCT)
        valentina_neto = sal_v_q + aux_v - ded_v

    tag = f"{p_ini.strftime('%Y%m%d')}_{p_fin.strftime('%Y%m%d')}"
    crear_reporte_horarios(resultados_t, p_ini, p_fin, f"REPORTE_HORARIOS_{tag}.xlsx")
    crear_colilla_pago(resultados_t, p_ini, p_fin, f"COLILLA_DE_PAGO_SOCCER7_{tag}.xlsx")
    gran_total = crear_resumen_nomina(resultados_t, valentina_neto, p_ini, p_fin, f"RESUMEN_NOMINA_{tag}.xlsx")

    print(f"\n{'Nombre':<30} {'H. Trab':>8} {'Tipo':>10} {'Noct $':>11} {'Neto a Pagar':>14}")
    print("─" * 80)
    for emp, t in resultados_t:
        datos = COLABORADORES.get(emp['nombre'], (emp['id'], emp['nombre'], "", "", "", "", SALARIO_MINIMO, "empleado"))
        tipo = datos[7]
        if tipo == "prestador":
            neto = t['val_total_prest']; tipo_label = "PRESTADOR"
        else:
            sal_q = t['salario'] / 2; days = t['dias_trab']
            aux = (AUXILIO_TRANSPORTE_ANUAL/2) * days / 15
            ibc = sal_q + t['val_en'] + t['val_noct']
            ded = ibc * (PENSION_PCT + SALUD_PCT)
            neto = sal_q + aux + t['val_en'] + t['val_noct'] - ded
            tipo_label = f"+{t['en_h']:.1f}h" if t['en_h'] > 0 else (f"DEBE {t['deu_h']:.0f}h" if t['deu_h'] > 0.1 else "✓")
        print(f"  {emp['nombre']:<28} {t['tot_h']:>7.1f}h {tipo_label:>10}  ${t['val_noct']:>9,.0f}  ${neto:>12,.0f}")
    if valentina_neto > 0:
        print(f"  {'VALENTINA GRANDA AGUDELO':<28} {'(admin)':>7}  {'sin reloj':>10}  ${'0':>9}  ${valentina_neto:>12,.0f}")
    print("─" * 80)
    print(f"  {'GRAN TOTAL NÓMINA':>52}             ${gran_total:>12,.0f}\n")


if __name__ == "__main__":
    main()
