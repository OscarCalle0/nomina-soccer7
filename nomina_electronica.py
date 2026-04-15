"""
nomina_electronica.py — Genera el Excel de Nómina Electrónica
Replica exactamente el formato NOMINA_ELECTRONICA_2026.xlsx por mes
con hojas ENE, FEB, MAR... y hoja RESUMEN EXTRAS
"""

import io
import calendar as cal_module
from datetime import datetime
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESES = {1:"ENE",2:"FEB",3:"MAR",4:"ABR",5:"MAY",6:"JUN",
         7:"JUL",8:"AGO",9:"SEP",10:"OCT",11:"NOV",12:"DIC"}

MESES_NOMBRE = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
                7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}

# Cuentas contables
COD_CUENTAS = {
    "SALARIO": "510506",
    "AUXILIO_TRANSPORTE": "510527",
    "H_EXTRAS": "510515",
    "PRIMA": "25101003",
    "RECARGO_NOCTURNO": "0",
    "SALUD": "237005",
    "PENSION": "238030",
    "PRESTAMOS": "0",
    "PROVISION_CESANTIAS": "0",
    "PROVISION_INT_CESANTIAS": "0",
    "PROVISION_VACACIONES": "0",
    "INCAPACIDAD": "0",
    "LICENCIA_NO_REM": "0",
    "VACACIONES": "",
}


def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(color): return PatternFill("solid", fgColor=color)
def _fnt(bold=False, color="000000", size=9): return Font(name="Arial", bold=bold, color=color, size=size)
def _aln(h="left", v="center"): return Alignment(horizontal=h, vertical=v)

def _sc(ws, r, c, v, bold=False, bg=None, fc="000000", ha="left", fmt=None, size=9):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _fnt(bold=bold, color=fc, size=size)
    cell.alignment = _aln(h=ha)
    if bg: cell.fill = _fill(bg)
    if fmt: cell.number_format = fmt
    return cell


def crear_hoja_mes(wb: openpyxl.Workbook, mes: int, anio: int,
                   quincenas_mes: List[Dict], colaboradores_mes: List[Dict]):
    """
    Crea o actualiza la hoja de un mes con el formato exacto del original.
    quincenas_mes: [dict_q1, dict_q2] — resultados de cada quincena
    colaboradores_mes: lista de Colaborador (como dict) activos en el mes
    """
    nombre_hoja = MESES[mes]
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)
    ws.sheet_view.showGridLines = False

    # Anchos de columna aproximados al original
    col_widths = [2, 28, 12, 14, 14, 14, 14, 2, 12, 18, 10, 2, 18, 10, 20, 10, 10, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ENCABEZADO ─────────────────────────────────────────────────────────────
    ws.merge_cells("B1:T1")
    ws["B1"].value = "GRANDA VARGAS S.A.S."
    ws["B1"].font = _fnt(bold=True, size=12)
    ws["B1"].alignment = _aln("center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("B2:T2")
    ws["B2"].value = f"NOMINA ELECTRONICA  {MESES_NOMBRE[mes]} {anio}"
    ws["B2"].font = _fnt(bold=True, size=11)
    ws["B2"].alignment = _aln("center")

    # Total nóminas
    total_nomina_mes = sum(c.get("total_nomina_mes", 0) for c in colaboradores_mes)
    ws.merge_cells("B3:G3")
    ws["B3"].value = "SALARIO MENSUAL"
    ws["B3"].font = _fnt(bold=True, size=9)
    ws["L3"].value = "TOTAL NOMINAS"
    ws["L3"].font = _fnt(bold=True, size=9)
    ws["P3"].value = total_nomina_mes
    ws["P3"].font = _fnt(bold=True, size=9)
    ws["P3"].number_format = "#,##0"

    ws.row_dimensions[3].height = 14

    # ── POR CADA COLABORADOR ───────────────────────────────────────────────────
    fila = 5
    for idx, col_data in enumerate(colaboradores_mes, 1):
        nombre = col_data.get("nombre_completo", "")
        q1 = col_data.get("q1", {})  # quincena del 15
        q2 = col_data.get("q2", {})  # quincena del 30/31

        # ── Cabecera del colaborador ──
        ws.merge_cells(f"B{fila}:F{fila}")
        ws[f"B{fila}"].value = nombre
        ws[f"B{fila}"].font = _fnt(bold=True, size=9)
        ws.cell(fila, 7).value = idx
        ws.cell(fila, 7).font = _fnt(bold=True, size=9)

        # Columnas resumen totales (días trabajados, etc.)
        ws.cell(fila, 8).value = "DIAS TRABAJADO"
        ws.cell(fila, 9).value = col_data.get("dias_trabajados", 30)
        ws.cell(fila, 10).value = "DIAS AUX TRANS"
        ws.cell(fila, 11).value = col_data.get("dias_aux_trans", 30)
        ws.cell(fila, 12).value = "HORAS EXTRAS"
        ws.cell(fila, 13).value = col_data.get("horas_extras_total", 0)
        ws.cell(fila, 14).value = "RECARGO NOCTURNO"
        ws.cell(fila, 15).value = col_data.get("recargo_noct_total_h", 0)
        for col_n in range(8, 16):
            ws.cell(fila, col_n).font = _fnt(size=8)
        # Notas
        notas = col_data.get("notas", "")
        if notas:
            ws.merge_cells(f"S{fila}:T{fila}")
            ws.cell(fila, 19).value = notas
            ws.cell(fila, 19).font = _fnt(size=8)
        fila += 1

        # ── Sub-cabecera quincenas ──
        sub_hdrs = ["INGRESO", "QUINCENA DEL 15", "HORAS", "", "QUINCENA DEL 30", "HORAS",
                    "TOTAL NOMINA DEL MES", "PRIMA", "VACACIONES", "CESANTIAS", "INT DE CESANTIAS", "NOTAS"]
        ws.cell(fila, 9).value = "INGRESO"
        ws.cell(fila, 10).value = "QUINCENA DEL 15"
        ws.cell(fila, 11).value = "HORAS"
        ws.cell(fila, 13).value = "QUINCENA DEL 30"
        ws.cell(fila, 14).value = "HORAS"
        ws.cell(fila, 15).value = "TOTAL NOMINA DEL MES"
        ws.cell(fila, 16).value = "PRIMA"
        ws.cell(fila, 17).value = "VACACIONES"
        ws.cell(fila, 18).value = "CESANTIAS"
        ws.cell(fila, 19).value = "INT DE CESANTIAS"
        ws.cell(fila, 20).value = "NOTAS"
        for col_n in range(9, 21):
            ws.cell(fila, col_n).font = _fnt(bold=True, size=8, color="1B3A5C")
            ws.cell(fila, col_n).fill = _fill("EEF4FB")
        fila += 1

        # ── Filas de conceptos ──
        def row_concepto(concepto_nom, cod_cuenta, valor_total,
                         v_q1, h_q1, v_q2, h_q2,
                         prima=0, vacac=0, ces=0, int_ces=0, nota=""):
            nonlocal fila
            _sc(ws, fila, 2, concepto_nom, size=9)
            _sc(ws, fila, 3, cod_cuenta, size=9)
            ws.cell(fila, 4).value = valor_total
            ws.cell(fila, 4).font = _fnt(size=9)
            ws.cell(fila, 4).number_format = "#,##0.##"
            # Q1
            ws.cell(fila, 10).value = v_q1
            ws.cell(fila, 10).font = _fnt(size=9)
            ws.cell(fila, 10).number_format = "#,##0.##"
            ws.cell(fila, 11).value = h_q1
            ws.cell(fila, 11).font = _fnt(size=9)
            ws.cell(fila, 11).number_format = "0.##"
            # Q2
            ws.cell(fila, 13).value = v_q2
            ws.cell(fila, 13).font = _fnt(size=9)
            ws.cell(fila, 13).number_format = "#,##0.##"
            ws.cell(fila, 14).value = h_q2
            ws.cell(fila, 14).font = _fnt(size=9)
            ws.cell(fila, 14).number_format = "0.##"
            # Total mes
            ws.cell(fila, 15).value = (v_q1 or 0) + (v_q2 or 0)
            ws.cell(fila, 15).font = _fnt(size=9)
            ws.cell(fila, 15).number_format = "#,##0.##"
            # Prestaciones
            for c_n, v in zip([16,17,18,19],[prima,vacac,ces,int_ces]):
                ws.cell(fila, c_n).value = v
                ws.cell(fila, c_n).font = _fnt(size=9)
                ws.cell(fila, c_n).number_format = "#,##0.##"
            if nota:
                ws.cell(fila, 20).value = nota
                ws.cell(fila, 20).font = _fnt(size=8)
            ws.row_dimensions[fila].height = 13
            fila += 1

        sal = col_data.get("salario_mensual", 0)
        sal_q = sal / 2
        dias_q1 = q1.get("dias_trab", 15)
        dias_q2 = q2.get("dias_trab", 15)
        aux_q   = 249095 / 2

        # SALARIO
        row_concepto("SALARIO", COD_CUENTAS["SALARIO"],
                     q1.get("sal_q", sal_q) + q2.get("sal_q", sal_q),
                     q1.get("sal_q", sal_q), dias_q1,
                     q2.get("sal_q", sal_q), dias_q2,
                     nota=col_data.get("notas_q1",""))

        # AUXILIO TRANSPORTE
        aux1 = aux_q * dias_q1 / 15
        aux2 = aux_q * dias_q2 / 15
        row_concepto("AUXILIO TRANSPORTE", COD_CUENTAS["AUXILIO_TRANSPORTE"],
                     aux1 + aux2, aux1, dias_q1, aux2, dias_q2)

        # H. EXTRAS
        ve1 = q1.get("val_en", 0) + q1.get("val_ee", 0)
        ve2 = q2.get("val_en", 0) + q2.get("val_ee", 0)
        he1 = q1.get("en_h", 0) + q1.get("ee_h", 0)
        he2 = q2.get("en_h", 0) + q2.get("ee_h", 0)
        row_concepto("H. EXTRAS", COD_CUENTAS["H_EXTRAS"],
                     ve1 + ve2, ve1 or None, he1 or None, ve2 or None, he2 or None)

        # PRIMA DE SERVICIOS
        row_concepto("PRIMA DE SERVICIOS", COD_CUENTAS["PRIMA"],
                     col_data.get("prima", 0), 0, 0, 0, 0)

        # RECARGO NOCTURNO
        vn1 = q1.get("val_noct", 0)
        vn2 = q2.get("val_noct", 0)
        hn1 = q1.get("noct_h", 0)
        hn2 = q2.get("noct_h", 0)
        row_concepto("RECARGO NOCTURNO", COD_CUENTAS["RECARGO_NOCTURNO"],
                     vn1 + vn2, vn1 or None, hn1 or None, vn2 or None, hn2 or None)

        # SALUD (negativo)
        ibc1 = q1.get("sal_q", sal_q) + q1.get("val_en",0) + q1.get("val_noct",0)
        ibc2 = q2.get("sal_q", sal_q) + q2.get("val_en",0) + q2.get("val_noct",0)
        salud1 = -ibc1 * 0.04
        salud2 = -ibc2 * 0.04
        row_concepto("SALUD", COD_CUENTAS["SALUD"],
                     salud1 + salud2, salud1, None, salud2, None)

        # PENSION (negativo)
        pen1 = -ibc1 * 0.04
        pen2 = -ibc2 * 0.04
        row_concepto("PENSION", COD_CUENTAS["PENSION"],
                     pen1 + pen2, pen1, None, pen2, None)

        # PRESTAMOS
        row_concepto("PRESTAMOS", COD_CUENTAS["PRESTAMOS"],
                     0, 0, None, 0, None)

        # VACACIONES (si hay)
        vac_val = col_data.get("vacaciones_val", 0)
        vac_dias = col_data.get("vacaciones_dias", 0)
        row_concepto("VACACIONES", COD_CUENTAS["VACACIONES"],
                     vac_val,
                     q1.get("vac_val", 0) or None, q1.get("vac_dias", 0) or None,
                     q2.get("vac_val", 0) or None, q2.get("vac_dias", 0) or None)

        # PROVISION CESANTIAS
        row_concepto("PROVISION CESANTIAS", COD_CUENTAS["PROVISION_CESANTIAS"],
                     col_data.get("prov_cesantias", 0), 0, None, 0, None)

        # PROVISION INT. CESANTIAS
        row_concepto("PROVISION INT. CESANTIAS", COD_CUENTAS["PROVISION_INT_CESANTIAS"],
                     col_data.get("prov_int_ces", 0), 0, None, 0, None)

        # PROVISION VACACIONES
        row_concepto("PROVISION VACACIONES", COD_CUENTAS["PROVISION_VACACIONES"],
                     col_data.get("prov_vacaciones", 0), 0, None, 0, None)

        # Novedades especiales (incapacidades, etc.)
        for nov_row in col_data.get("novedades_nomina", []):
            row_concepto(nov_row["desc"], "0",
                         nov_row.get("valor_total", 0),
                         nov_row.get("val_q1", 0) or None, nov_row.get("h_q1") or None,
                         nov_row.get("val_q2", 0) or None, nov_row.get("h_q2") or None)

        # ── FILA TOTALES ──────────────────────────────────────────────────────
        tot_q1  = (q1.get("sal_q", sal_q) + aux1 + ve1 + vn1
                   + salud1 + pen1 + q1.get("nov_devengado",0) - q1.get("nov_deduccion",0))
        tot_q2  = (q2.get("sal_q", sal_q) + aux2 + ve2 + vn2
                   + salud2 + pen2 + q2.get("nov_devengado",0) - q2.get("nov_deduccion",0))
        tot_mes = tot_q1 + tot_q2

        _sc(ws, fila, 2, "TOTAL", bold=True, size=9)
        ws.cell(fila, 9).value  = "TOTALES"
        ws.cell(fila, 9).font   = _fnt(bold=True, size=9)
        ws.cell(fila, 10).value = round(tot_q1)
        ws.cell(fila, 10).font  = _fnt(bold=True, size=9)
        ws.cell(fila, 10).number_format = "#,##0"
        ws.cell(fila, 13).value = round(tot_q2)
        ws.cell(fila, 13).font  = _fnt(bold=True, size=9)
        ws.cell(fila, 13).number_format = "#,##0"
        ws.cell(fila, 15).value = round(tot_mes)
        ws.cell(fila, 15).font  = _fnt(bold=True, size=9)
        ws.cell(fila, 15).number_format = "#,##0"
        for col_n in range(2, 21):
            ws.cell(fila, col_n).fill = _fill("EEF4FB")
        ws.row_dimensions[fila].height = 14; fila += 1

        # TOTAL NOMINA DEL MES
        ws.cell(fila, 9).value  = "TOTAL NOMINA DEL MES"
        ws.cell(fila, 9).font   = _fnt(bold=True, size=9)
        ws.cell(fila, 15).value = round(tot_mes)
        ws.cell(fila, 15).font  = _fnt(bold=True, size=9)
        ws.cell(fila, 15).number_format = "#,##0"
        ws.row_dimensions[fila].height = 13; fila += 2

    # ── TOTALES DEL MES ────────────────────────────────────────────────────────
    total_q1 = sum(c.get("tot_q1", 0) for c in colaboradores_mes)
    total_q2 = sum(c.get("tot_q2", 0) for c in colaboradores_mes)

    ws.cell(fila, 3).value = "QUINCENA 1"
    ws.cell(fila, 3).font  = _fnt(bold=True)
    ws.cell(fila, 4).value = round(total_q1)
    ws.cell(fila, 4).font  = _fnt(bold=True)
    ws.cell(fila, 4).number_format = "#,##0"
    fila += 1

    ws.cell(fila, 3).value = "QUINCENA 2"
    ws.cell(fila, 3).font  = _fnt(bold=True)
    ws.cell(fila, 4).value = round(total_q2)
    ws.cell(fila, 4).font  = _fnt(bold=True)
    ws.cell(fila, 4).number_format = "#,##0"
    fila += 1

    ws.cell(fila, 9).value  = "TOTAL NOMINA DEL MES"
    ws.cell(fila, 9).font   = _fnt(bold=True)
    ws.cell(fila, 15).value = round(total_q1 + total_q2)
    ws.cell(fila, 15).font  = _fnt(bold=True)
    ws.cell(fila, 15).number_format = "#,##0"
    fila += 1

    ws.cell(fila, 3).value = "TOTAL NOMINA"
    ws.cell(fila, 3).font  = _fnt(bold=True)
    ws.cell(fila, 4).value = round(total_q1 + total_q2)
    ws.cell(fila, 4).font  = _fnt(bold=True)
    ws.cell(fila, 4).number_format = "#,##0"
    fila += 1

    ws.cell(fila, 3).value = "DIFERENCIA"
    ws.cell(fila, 3).font  = _fnt(bold=True)
    ws.cell(fila, 4).value = -(total_q1 + total_q2)
    ws.cell(fila, 4).font  = _fnt(bold=True)
    ws.cell(fila, 4).number_format = "#,##0"
    fila += 1

    primas_mes = sum(c.get("prima", 0) for c in colaboradores_mes)
    vac_mes    = sum(c.get("vacaciones_val", 0) for c in colaboradores_mes)
    ws.cell(fila, 3).value = "PRIMAS"
    ws.cell(fila, 4).value = round(primas_mes)
    ws.cell(fila, 4).number_format = "#,##0"
    fila += 1
    ws.cell(fila, 3).value = "VACACIONES"
    ws.cell(fila, 4).value = round(vac_mes)
    ws.cell(fila, 4).number_format = "#,##0"

    ws.freeze_panes = "B6"


def generar_nomina_electronica_xlsx(meses_data: Dict[int, Dict], anio: int) -> bytes:
    """
    Genera el Excel completo de nómina electrónica.
    meses_data: {mes_numero: {"colaboradores": [...], "quincenas": [...]}}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Crear hojas por mes
    for mes in sorted(meses_data.keys()):
        data = meses_data[mes]
        crear_hoja_mes(wb, mes, anio,
                       quincenas_mes=data.get("quincenas", []),
                       colaboradores_mes=data.get("colaboradores", []))

    # Hoja RESUMEN EXTRAS (si hay datos)
    ws_res = wb.create_sheet("RESUMEN EXTRAS")
    ws_res["B1"].value = "LIQUIDACION PARA CESANTIAS"
    ws_res["B1"].font = _fnt(bold=True, size=10)

    # Hoja LICENCIA
    ws_lic = wb.create_sheet("LICENCIA")
    ws_lic["B1"].value = "LICENCIA NO REMUNERADA"
    ws_lic["B1"].font = _fnt(bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def preparar_datos_mes_desde_historico(historico: List[Dict], mes: int, anio: int,
                                        colaboradores: List) -> Dict:
    """
    Prepara los datos de un mes combinando las dos quincenas del histórico.
    """
    # Filtrar quincenas del mes
    q_mes = []
    for h in historico:
        try:
            ini = datetime.fromisoformat(h.get("periodo_ini", ""))
            if ini.month == mes and ini.year == anio:
                q_mes.append(h)
        except:
            pass

    q_mes.sort(key=lambda x: x.get("periodo_ini", ""))
    q1_data = q_mes[0] if len(q_mes) > 0 else {}
    q2_data = q_mes[1] if len(q_mes) > 1 else {}

    # Construir datos por colaborador
    col_dict = {}
    for col in colaboradores:
        clave = col.nombre_reloj if hasattr(col, 'nombre_reloj') else col.get("nombre_reloj", "")
        nombre_c = col.nombre_completo if hasattr(col, 'nombre_completo') else col.get("nombre_completo", "")
        col_dict[clave] = {
            "nombre_completo": nombre_c,
            "salario_mensual": col.salario_mensual if hasattr(col, 'salario_mensual') else col.get("salario_mensual", 0),
            "tipo": col.tipo if hasattr(col, 'tipo') else col.get("tipo", "empleado"),
            "q1": {}, "q2": {},
            "notas_q1": "", "notas_q2": "",
            "prima": 0, "vacaciones_val": 0, "vacaciones_dias": 0,
            "prov_cesantias": 0, "prov_int_ces": 0, "prov_vacaciones": 0,
            "novedades_nomina": [],
        }

    # Llenar con datos de quincenas
    for q_idx, q_data in enumerate([(q1_data, "q1"), (q2_data, "q2")]):
        q, qkey = q_data
        for col_res in q.get("colaboradores", []):
            nombre = col_res.get("nombre", "")
            if nombre in col_dict:
                d = col_dict[nombre]
                sal_q = d["salario_mensual"] / 2
                dias_trab = col_res.get("dias_trab", 15)
                d[qkey] = {
                    "sal_q":        sal_q * dias_trab / 15,
                    "dias_trab":    dias_trab,
                    "val_en":       col_res.get("val_en", 0),
                    "val_ee":       col_res.get("val_ee", 0),
                    "val_noct":     col_res.get("val_noct", 0),
                    "noct_h":       col_res.get("noct_h", 0),
                    "en_h":         col_res.get("en_h", 0),
                    "ee_h":         col_res.get("ee_h", 0),
                    "nov_devengado":col_res.get("nov_devengado", 0),
                    "nov_deduccion":col_res.get("nov_deduccion", 0),
                }
                # Novedades
                notas = col_res.get("novedades_desc", "")
                if notas: d[f"notas_{qkey}"] = notas

                # Acumular totales
                ibc = d[qkey]["sal_q"] + d[qkey]["val_en"] + d[qkey]["val_noct"]
                aux = (249095/2) * dias_trab / 15
                tot = (d[qkey]["sal_q"] + aux + d[qkey]["val_en"] + d[qkey]["val_noct"]
                       - ibc * 0.04 - ibc * 0.04
                       + d[qkey]["nov_devengado"] - d[qkey]["nov_deduccion"])
                d[f"tot_{qkey}"] = tot

    # Calcular totales del mes
    for nombre, d in col_dict.items():
        d["dias_trabajados"]  = d["q1"].get("dias_trab",0) + d["q2"].get("dias_trab",0)
        d["dias_aux_trans"]   = d["dias_trabajados"]
        d["horas_extras_total"] = (d["q1"].get("en_h",0)+d["q1"].get("ee_h",0)+
                                    d["q2"].get("en_h",0)+d["q2"].get("ee_h",0))
        d["recargo_noct_total_h"] = d["q1"].get("noct_h",0) + d["q2"].get("noct_h",0)
        d["total_nomina_mes"] = d.get("tot_q1",0) + d.get("tot_q2",0)
        d["notas"] = " / ".join(filter(None,[d.get("notas_q1",""), d.get("notas_q2","")]))

    return {
        "colaboradores": list(col_dict.values()),
        "quincenas": [q1_data, q2_data],
    }
